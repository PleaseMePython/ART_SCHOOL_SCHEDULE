"""Алгоритм преобразования выходного файла в Excel."""

from pathlib import Path
from typing import List

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

import src.parsers.schedule as schedule
from src.parsers import NodeType


class Excel:
    """Преобразователь в Excel."""

    def process(self, template_path: Path, src_path: Path, dst_path: Path):
        """Преобразование файла.

        :arg template_path: Путь к шаблону
        :arg src_path: Путь к преобразуемому файлу
        :arg dst_path: Путь для сохранения результата
        """
        # Читаем XML с расписанием
        with Path.open(src_path, 'r', encoding='utf-8') as file:
            src_xml = file.read()
        s: schedule.Schedule = schedule.parse_schedule(src_xml)

        if not s:
            Exception('Исходный файл имеет неверный формат')

        wb = load_workbook(template_path)
        global_names = wb.defined_names
        wsp = wb['Шаблон']
        for teacher in s.teachers:
            wst = wb.copy_worksheet(wsp)
            wst.title = teacher.name
            for gn, gn_data in global_names.items():
                gn_ref = gn_data.attr_text.replace('Шаблон', f"'{teacher.name}'", 1)
                loc_name = DefinedName(gn, attr_text=gn_ref)
                wst.defined_names.add(loc_name)

            self.__model_to_sheet(ws=wst, model=teacher)

        wb.remove(wsp)
        wb.save(dst_path)

    def __fill_range(
        self,
        ws: Worksheet,
        name: str,
        value: str,
        row: int = 1,
        col: int = 1,
    ):
        """Заполнение именованного диапазона ячеек значением.

        :arg ws: Лист
        :arg name: Имя диапазона
        :arg value: Записываемое значение
        :arg row: Строка внутри диапазона
        :arg col: Столбец внутри диапазона
        """
        dest = ws.defined_names[name].destinations
        for _, r in dest:
            wr = ws[r]
            if not isinstance(wr, tuple):
                ws[r] = value
            else:
                wr_row = wr[row - 1]
                wr_row[col - 1].value = value

    def __model_to_sheet(
        self,
        /,
        ws: Worksheet,
        model: schedule.ScheduleBaseModel | List,
        parent_descr: str = '',
    ):
        """Рекурсивный вывод модели на лист.

        :arg ws: Лист
        :arg model: Модель
        :arg parent_descr: Имя узла модели верхнего уровня
        """
        if isinstance(model, list):
            for lesson in model:
                self.__fill_range(
                    ws=ws,
                    name=parent_descr,
                    value=lesson.group,
                    row=lesson.npp,
                    col=1,
                )
                self.__fill_range(
                    ws=ws,
                    name=parent_descr,
                    value=lesson.subject,
                    row=lesson.npp,
                    col=2,
                )
        if not isinstance(model, schedule.ScheduleBaseModel):
            return
        schedule_model: schedule.ScheduleBaseModel = model
        for model_fields_name in schedule_model.model_fields_set:
            descr = parent_descr + schedule_model.model_fields[model_fields_name].path
            if (
                schedule_model.model_fields[model_fields_name].location
                == NodeType.ATTRIBUTE
            ):
                self.__fill_range(
                    ws=ws,
                    name=descr,
                    value=getattr(schedule_model, model_fields_name),
                )
            else:
                child_model = getattr(schedule_model, model_fields_name)
                self.__model_to_sheet(ws=ws, model=child_model, parent_descr=descr)
